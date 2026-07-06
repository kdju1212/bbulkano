from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from services.formula_eval import effective_value
from utils.range_utils import RangeSpec, col_index, parse_range, split_sheet_ref


class ExcelService:
    def __init__(self, template_file: Path) -> None:
        self.workbook = load_workbook(template_file, keep_vba=True, data_only=False)
        self.workbook.calculation.fullCalcOnLoad = True
        self.workbook.calculation.forceFullCalc = True
        self.workbook.calculation.calcMode = "auto"

    def save(self, output_file: Path) -> None:
        output_file.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_file)

    def write_raw_sheet(self, sheet_name: str, rows: list[list[Any]]) -> None:
        title = sanitize_sheet_name(sheet_name, taken=set(self.workbook.sheetnames))
        ws = self.workbook.create_sheet(title=title)
        for row_values in rows:
            ws.append([normalize_value(value) for value in row_values])

    def clear_groups(self, rule: dict[str, Any]) -> None:
        ws = self.workbook[rule["sheet"]]
        for group in rule["groups"]:
            spec = parse_range(group["range"])
            start_row = group["start_row"]
            for row in range(start_row, ws.max_row + 1):
                for col in range(spec.min_col, spec.max_col + 1):
                    ws.cell(row=row, column=col).value = None

    def write_block(
        self,
        sheet_name: str,
        start_row: int,
        start_col: int,
        values: list[list[Any]],
        style_row: int | None = None,
    ) -> int:
        ws = self.workbook[sheet_name]
        if not values:
            return start_row - 1
        for row_offset, row_values in enumerate(values):
            row_num = start_row + row_offset
            for col_offset, value in enumerate(row_values):
                col_num = start_col + col_offset
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = normalize_value(value)
                if style_row:
                    self.copy_cell_style(ws.cell(row=style_row, column=col_num), cell)
        return start_row + len(values) - 1

    def append_row_for(self, sheet_name: str, ref_col: int, minimum_row: int = 6) -> int:
        ws = self.workbook[sheet_name]
        for row in range(ws.max_row, minimum_row - 1, -1):
            if ws.cell(row=row, column=ref_col).value not in (None, ""):
                return row + 1
        return minimum_row

    def last_data_row(self, sheet_name: str, ref_column: str | int, start_row: int = 1) -> int:
        ws = self.workbook[sheet_name]
        col = col_index(ref_column)
        for row in range(ws.max_row, start_row - 1, -1):
            if effective_value(ws, row, col) not in (None, ""):
                return row
        return start_row - 1

    def filldown(self, rule: dict[str, Any]) -> None:
        ws = self.workbook[rule["target_sheet"]]
        _, ref = split_sheet_ref(rule["ref_column"])
        ref_col = col_index(ref.rstrip("0123456789"))
        top = parse_range(rule["range_top"])
        source_row = top.min_row or 1
        last_row = self.last_data_row(ws.title, ref_col, source_row)
        if last_row <= source_row:
            return
        for target_row in range(source_row + 1, last_row + 1):
            for col in range(top.min_col, top.max_col + 1):
                source = ws.cell(row=source_row, column=col)
                target = ws.cell(row=target_row, column=col)
                self.copy_cell_style(source, target)
                if isinstance(source.value, str) and source.value.startswith("="):
                    origin = f"{get_column_letter(col)}{source_row}"
                    target.value = Translator(source.value, origin=origin).translate_formula(
                        f"{get_column_letter(col)}{target_row}"
                    )
                else:
                    target.value = source.value

    def copy_mapped_rows(
        self,
        source_sheet: str,
        target_sheet: str,
        mappings: list[list[str]],
        markings: list[list[str]] | None = None,
        start_row: int | None = None,
        link_formulas: bool = False,
    ) -> int:
        src_ws = self.workbook[source_sheet]
        dst_ws = self.workbook[target_sheet]
        source_specs = [parse_range(src) for src, _ in mappings]
        dest_specs = [parse_range(dst) for _, dst in mappings]
        first_source_row = source_specs[0].min_row or 1
        last_source_row = self._last_row_from_specs(src_ws, source_specs, first_source_row)
        if last_source_row < first_source_row:
            return start_row or self.append_row_for(target_sheet, dest_specs[0].min_col)
        dest_row = start_row or self._dest_start_row(dst_ws, dest_specs[0])
        row_count = last_source_row - first_source_row + 1
        for offset in range(row_count):
            for source_spec, dest_spec in zip(source_specs, dest_specs):
                src_row = first_source_row + offset
                dst_row = dest_row + offset
                for col_offset in range(source_spec.width):
                    src_col = source_spec.min_col + col_offset
                    dst_col = dest_spec.min_col + col_offset
                    src_cell = src_ws.cell(row=src_row, column=src_col)
                    dst_cell = dst_ws.cell(row=dst_row, column=dst_col)
                    if link_formulas and isinstance(src_cell.value, str) and src_cell.value.startswith("="):
                        dst_cell.value = f"='{source_sheet}'!{src_cell.coordinate}"
                    else:
                        dst_cell.value = src_cell.value
                    self.copy_cell_style(dst_ws.cell(row=max(dest_spec.min_row or 6, 1), column=dst_col), dst_cell)
            for col, value in markings or []:
                dst_ws.cell(row=dest_row + offset, column=col_index(col)).value = value
        return dest_row + row_count - 1

    def apply_rules(self, rule_group: dict[str, Any]) -> None:
        ws = self.workbook[rule_group["target_sheet"]]
        ref_col = col_index(rule_group.get("ref_column", "A").split("!")[-1].rstrip("0123456789"))
        last_row = self.last_data_row(ws.title, ref_col, 1)
        for row in range(1, last_row + 1):
            for rule in rule_group["rules"]:
                if self._matches(ws, row, rule):
                    action = rule["action"]
                    ws.cell(row=row, column=col_index(action["column"])).value = action["value"]

    def filter_copy(self, rule: dict[str, Any]) -> int:
        src_ws = self.workbook[rule["source_sheet"]]
        dst_ws = self.workbook[rule["target_sheet"]]
        mappings = [(parse_range(src), parse_range(dst)) for src, dst in rule["map"]]
        start_row = 6
        last_row = self._last_row_from_specs(src_ws, [m[0] for m in mappings], start_row)
        dest_row = self.append_row_for(dst_ws.title, mappings[0][1].min_col)
        copied = 0
        for row in range(start_row, last_row + 1):
            if not self._row_matches_conditions(src_ws, row, rule["filter_conditions"]):
                continue
            for source_spec, dest_spec in mappings:
                for col_offset in range(source_spec.width):
                    value = effective_value(src_ws, row, source_spec.min_col + col_offset)
                    target = dst_ws.cell(row=dest_row + copied, column=dest_spec.min_col + col_offset)
                    target.value = value
            copied += 1
        return copied

    def replace_in_column(self, sheet_name: str, column: str, replacements: list[dict[str, str]]) -> None:
        ws = self.workbook[sheet_name]
        col = col_index(column)
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str):
                for replacement in replacements:
                    value = value.replace(replacement["find"], replacement["replace"])
                ws.cell(row=row, column=col).value = value

    def format_date_column(self, sheet_name: str, column: str, number_format: str) -> None:
        ws = self.workbook[sheet_name]
        col = col_index(column)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str):
                parsed = parse_date(cell.value)
                if parsed:
                    cell.value = parsed
            if isinstance(cell.value, datetime):
                cell.number_format = number_format

    @staticmethod
    def copy_cell_style(source, target) -> None:
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format

    def _matches(self, ws: Worksheet, row: int, rule: dict[str, Any]) -> bool:
        if rule["type"] == "simple":
            return _match_one(ws, row, rule["match"])
        if rule["type"] == "multi":
            return all(_match_one(ws, row, condition) for condition in rule["match"])
        raise ValueError(f"지원하지 않는 rule type: {rule['type']}")

    def _row_matches_conditions(self, ws: Worksheet, row: int, conditions: list[dict[str, Any]]) -> bool:
        return all(_match_one(ws, row, condition) for condition in conditions)

    def _last_row_from_specs(self, ws: Worksheet, specs: list[RangeSpec], start_row: int) -> int:
        max_row = start_row - 1
        for spec in specs:
            for row in range(ws.max_row, start_row - 1, -1):
                if any(effective_value(ws, row, col) not in (None, "") for col in range(spec.min_col, spec.max_col + 1)):
                    max_row = max(max_row, row)
                    break
        return max_row

    def _dest_start_row(self, ws: Worksheet, spec: RangeSpec) -> int:
        if spec.min_row:
            return spec.min_row
        return self.append_row_for(ws.title, spec.min_col)


def _match_one(ws: Worksheet, row: int, condition: dict[str, Any]) -> bool:
    value = effective_value(ws, row, col_index(condition["column"]))
    text = "" if value is None else str(value)
    if condition.get("case_insensitive"):
        text_cmp = text.lower()
        values = [str(v).lower() for v in condition.get("values", [])]
        contains_any = [str(v).lower() for v in condition.get("contains_any", [])]
    else:
        text_cmp = text
        values = [str(v) for v in condition.get("values", [])]
        contains_any = [str(v) for v in condition.get("contains_any", [])]
    if "values" in condition and text_cmp not in values:
        return False
    if "contains_any" in condition and not any(item in text_cmp for item in contains_any):
        return False
    return True


def normalize_value(value: Any) -> Any:
    if value == "":
        return None
    return value


def parse_date(value: str) -> datetime | None:
    text = value.strip()
    if text.endswith("-"):
        text = text[:-1]
    for fmt in ("%Y-%m-%d", "%Y.%m.%d.", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


_INVALID_SHEET_CHARS = str.maketrans({c: "_" for c in "\\/*?:[]"})


def sanitize_sheet_name(name: str, taken: set[str]) -> str:
    cleaned = name.translate(_INVALID_SHEET_CHARS).strip("'") or "Sheet"
    candidate = cleaned[:31]
    suffix = 2
    while candidate in taken:
        trimmed = cleaned[: 31 - len(f"_{suffix}")]
        candidate = f"{trimmed}_{suffix}"
        suffix += 1
    return candidate
